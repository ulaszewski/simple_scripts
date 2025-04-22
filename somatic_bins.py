import argparse
import openpyxl
from openpyxl.utils import get_column_letter

# Read data from input file
def read_data(file_path):
    data = []
    with open(file_path, 'r') as file:
        for line in file:
            parts = line.strip().split()
            if len(parts) == 3:
                category = parts[0]
                number = int(parts[1])
                value = int(parts[2])
                data.append((category, number, value))
    return data

# Read max value per category from separate file
def read_category_max_values(file_path):
    limits = {}
    with open(file_path, 'r') as file:
        for line in file:
            parts = line.strip().split()
            if len(parts) == 2:
                category = parts[0]
                max_val = int(parts[1])
                limits[category] = max_val
    return limits

# Count occurrences per bin where value > 0
def count_occurrences_by_category(data, category_max_values, interval_size):
    results = {}

    for category, max_val in category_max_values.items():
        interval_counts = {i: 0 for i in range(0, max_val, interval_size)}
        filtered_data = [entry for entry in data if entry[0] == category]
        for _, number, value in filtered_data:
            if value > 0 and 0 <= number < max_val:
                interval = (number // interval_size) * interval_size
                if interval in interval_counts:
                    interval_counts[interval] += 1
        results[category] = interval_counts

    return results

# Save results to Excel file
def save_to_excel(results, interval_size, output_file):
    # Collect all unique bins
    all_intervals = set()
    for intervals in results.values():
        all_intervals.update(intervals.keys())

    all_intervals = sorted(all_intervals)
    interval_labels = [f"{start}-{start + interval_size - 1}" for start in all_intervals]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    # Write header
    ws.cell(row=1, column=1, value="")
    for col_idx, label in enumerate(interval_labels, start=2):
        ws.cell(row=1, column=col_idx, value=label)

    # Write category rows
    for row_idx, category in enumerate(sorted(results.keys()), start=2):
        ws.cell(row=row_idx, column=1, value=category)
        for col_idx, start in enumerate(all_intervals, start=2):
            count = results[category].get(start, "")
            ws.cell(row=row_idx, column=col_idx, value=count)

    # Auto adjust column widths
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    wb.save(output_file)

# Main function
def main():
    parser = argparse.ArgumentParser(description="Count interval occurrences per category and export to Excel.")
    parser.add_argument("input_file", help="Path to the input data file")
    parser.add_argument("category_file", help="Path to the category max value file")
    parser.add_argument("output_file", help="Path to the output Excel file (.xlsx)")
    parser.add_argument("--interval_size", type=int, default=100000, help="Size of each interval")
    args = parser.parse_args()

    data = read_data(args.input_file)
    category_max_values = read_category_max_values(args.category_file)
    results = count_occurrences_by_category(data, category_max_values, args.interval_size)
    save_to_excel(results, args.interval_size, args.output_file)

if __name__ == "__main__":
    main()
